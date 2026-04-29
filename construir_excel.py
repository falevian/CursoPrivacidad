"""
Construye un Excel multi-hoja con todos los datasets para uso fácil en clase.
Sólo incluye una muestra del dataset de movilidad (es muy grande para Excel).
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

OUT_DIR = "/home/claude/datos_curso"

print("Cargando CSVs...")
df_main = pd.read_csv(f"{OUT_DIR}/01_dataset_principal.csv", dtype={"codigo_postal": str, "dni": str})
df_hosp = pd.read_csv(f"{OUT_DIR}/02_dataset_hospital.csv", dtype={"codigo_postal": str})
df_social = pd.read_csv(f"{OUT_DIR}/03_dataset_red_social.csv")
df_mov = pd.read_csv(f"{OUT_DIR}/04_dataset_movilidad.csv")

# Muestra de movilidad para que quepa en Excel y se pueda inspeccionar
df_mov_sample = df_mov.head(50000).copy()

print("Creando workbook...")
wb = Workbook()

# ===== Hoja Guía =====
ws = wb.active
ws.title = "Guía"

guia = [
    ("Curso de Privacidad y Anonimización — UC3M", "title"),
    ("Datasets sintéticos para prácticas (3 h)", "subtitle"),
    ("", ""),
    ("Hoja", "Descripción"),
    ("01_principal", "Base de clientes (5.000 filas, 43 cols). Identificadores directos + cuasi-id + sensibles. Para entender qué hay que proteger."),
    ("02_hospital", "2.500 pacientes con (edad, sexo, CP) + diagnóstico. Para k-anonimidad y l-diversidad."),
    ("03_red_social", "3.750 perfiles públicos con valoraciones de películas. Para ataques de vinculabilidad estilo Netflix-IMDB."),
    ("04_movilidad_sample", "50.000 eventos de movilidad (CDR). Muestra del fichero completo de 251K filas. Para ataques de unicidad estilo Montjoye."),
    ("SOLUCIÓN", "Sólo profesor. Mapeo username → id_cliente para verificar ataques de vinculabilidad."),
    ("Diccionario", "Diccionario de variables del dataset principal con clasificación por sensibilidad (Art. 9 RGPD)."),
    ("", ""),
    ("Tipos de variables (slide del curso)", "section"),
    ("Identificadores directos", "id_cliente, dni, nombre, apellido1, apellido2, telefono, email, iban, direccion, usuario"),
    ("Cuasi-identificadores", "fecha_nacimiento, edad, sexo, codigo_postal, ciudad, provincia, nacionalidad"),
    ("Sensibles (Art. 9 RGPD)", "religion, orientacion_sexual, diabetes, hipertension, fumador, ansiedad, depresion, etc."),
    ("No sensibles", "estado_civil, num_hijos, altura, peso, imc, nivel_educativo, ocupacion, salario, etc."),
    ("", ""),
    ("Verificaciones automáticas tras la generación", "section"),
    ("k-anonimidad inicial (hospital, QID=edad+sexo+CP)", "k=1 al 100% — cada paciente es único antes de proteger"),
    ("Unicidad con 4 puntos (movilidad)", "97,3% — replica Montjoye et al. 2013 (95%)"),
    ("Distribución alturas H/M", "173,1 / 160,3 cm — coincide con ENSE 2017"),
    ("Top nombres", "Antonio, José, Manuel, María Carmen, María — coincide con INE"),
]

# Estilos
font_title = Font(name="Arial", size=18, bold=True, color="FFFFFF")
font_subtitle = Font(name="Arial", size=12, italic=True, color="FFFFFF")
font_section = Font(name="Arial", size=12, bold=True, color="FFFFFF")
font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
font_normal = Font(name="Arial", size=10)
fill_title = PatternFill("solid", start_color="1F4E78")
fill_section = PatternFill("solid", start_color="2E75B6")
fill_header = PatternFill("solid", start_color="5B9BD5")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for i, (a, b) in enumerate(guia, start=1):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    if b == "title":
        ws.cell(row=i, column=1).font = font_title
        ws.cell(row=i, column=1).fill = fill_title
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    elif b == "subtitle":
        ws.cell(row=i, column=1).font = font_subtitle
        ws.cell(row=i, column=1).fill = fill_title
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    elif b == "section":
        ws.cell(row=i, column=1).font = font_section
        ws.cell(row=i, column=1).fill = fill_section
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    elif a == "Hoja":
        ws.cell(row=i, column=1).font = font_header
        ws.cell(row=i, column=2).font = font_header
        ws.cell(row=i, column=1).fill = fill_header
        ws.cell(row=i, column=2).fill = fill_header
    else:
        ws.cell(row=i, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=i, column=2).font = font_normal
    ws.cell(row=i, column=1).alignment = align_left
    ws.cell(row=i, column=2).alignment = align_left

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 95
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 22

# ===== Hoja Diccionario de variables =====
ws_dict = wb.create_sheet("Diccionario")
diccionario = [
    ("Variable", "Tipo", "Sensibilidad", "Descripción"),
    ("id_cliente", "string", "Identificador directo", "ID interno del cliente (C000001…)"),
    ("dni", "string", "Identificador directo", "DNI español con letra de control válida"),
    ("nombre", "string", "Identificador directo", "Nombre (frecuencia INE)"),
    ("apellido1", "string", "Identificador directo", "Primer apellido (frecuencia INE)"),
    ("apellido2", "string", "Identificador directo", "Segundo apellido"),
    ("sexo", "categórica (M/F)", "Cuasi-identificador", "Sexo biológico"),
    ("fecha_nacimiento", "fecha", "Cuasi-identificador", "Fecha de nacimiento"),
    ("edad", "numérica", "Cuasi-identificador", "Edad en años"),
    ("nacionalidad", "categórica", "Cuasi-identificador", "Nacionalidad (distrib. INE)"),
    ("estado_civil", "categórica", "No sensible", "Estado civil"),
    ("num_hijos", "numérica", "No sensible", "Número de hijos"),
    ("telefono", "string", "Identificador directo", "Teléfono (móvil/fijo ES)"),
    ("email", "string", "Identificador directo", "Email plausible"),
    ("usuario", "string", "Identificador directo", "Nombre de usuario derivado del email"),
    ("password_hash_sha256", "string", "Sensible", "Hash de contraseña simulado"),
    ("direccion", "string", "Identificador directo", "Dirección postal"),
    ("codigo_postal", "string", "Cuasi-identificador", "CP coherente con provincia"),
    ("ciudad", "string", "Cuasi-identificador", "Ciudad"),
    ("provincia", "string", "Cuasi-identificador", "Provincia"),
    ("pais", "string", "Cuasi-identificador", "País"),
    ("altura_cm", "numérica", "No sensible", "Altura cm. Distrib. real ENSE 2017"),
    ("peso_kg", "numérica", "No sensible", "Peso kg. IMC realista por edad/sexo"),
    ("imc", "numérica", "No sensible", "Índice de masa corporal"),
    ("nivel_educativo", "categórica", "No sensible", "Distrib. INE 25-64 años"),
    ("situacion_laboral", "categórica", "No sensible", "Situación laboral"),
    ("ocupacion", "categórica", "No sensible", "Categoría ocupacional (CNO)"),
    ("salario_anual_euros", "numérica", "Sensible", "Salario bruto anual €. Correlacionado con todo"),
    ("iban", "string", "Identificador directo", "IBAN español con dígito de control válido"),
    ("tiene_hipoteca", "binaria", "No sensible", "Indicador hipoteca"),
    ("deuda_total_euros", "numérica", "Sensible", "Deuda total"),
    ("tiene_coche", "binaria", "No sensible", "Indicador vehículo"),
    ("religion", "categórica", "Sensible (Art. 9 RGPD)", "Religión declarada"),
    ("orientacion_sexual", "categórica", "Sensible (Art. 9 RGPD)", "Orientación sexual"),
    ("diabetes", "binaria", "Sensible (salud)", "Diagnóstico diabetes"),
    ("hipertension", "binaria", "Sensible (salud)", "Diagnóstico hipertensión"),
    ("colesterol_mgdl", "numérica", "Sensible (salud)", "Colesterol total"),
    ("fumador", "binaria", "Sensible", "Hábito tabáquico"),
    ("consumo_alcohol_ud_semana", "numérica", "Sensible", "Unidades alcohol/semana"),
    ("ansiedad_diagnosticada", "binaria", "Sensible (salud mental)", "Diagnóstico ansiedad"),
    ("depresion_diagnosticada", "binaria", "Sensible (salud mental)", "Diagnóstico depresión"),
    ("horas_internet_dia", "numérica", "No sensible", "Horas conectado al día"),
    ("gasto_mensual_online_euros", "numérica", "No sensible", "Gasto online mensual"),
    ("fecha_alta_cliente", "fecha", "No sensible", "Fecha de alta como cliente"),
]
for r, fila in enumerate(diccionario, start=1):
    for c, val in enumerate(fila, start=1):
        cell = ws_dict.cell(row=r, column=c, value=val)
        cell.font = font_normal if r > 1 else font_header
        cell.alignment = align_left
        if r == 1:
            cell.fill = fill_header
ws_dict.column_dimensions["A"].width = 30
ws_dict.column_dimensions["B"].width = 18
ws_dict.column_dimensions["C"].width = 26
ws_dict.column_dimensions["D"].width = 60

# ===== Función para volcar DataFrame a hoja =====
def df_a_hoja(wb, df, sheet_name, max_col_width=25):
    ws = wb.create_sheet(sheet_name)
    # Escribir cabeceras
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Escribir datos
    for r, fila in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(fila, start=1):
            ws.cell(row=r, column=c, value=val)
    # Ajustar columnas
    for c, col in enumerate(df.columns, start=1):
        letra = get_column_letter(c)
        ancho = min(max(len(str(col)), 12), max_col_width)
        ws.column_dimensions[letra].width = ancho
    # Congelar fila de cabecera
    ws.freeze_panes = "A2"

print("Volcando hojas...")
df_a_hoja(wb, df_main, "01_principal")
df_a_hoja(wb, df_hosp, "02_hospital")
df_a_hoja(wb, df_social, "03_red_social", max_col_width=40)
df_a_hoja(wb, df_mov_sample, "04_movilidad_sample")

# Solución
df_sol = pd.read_csv(f"{OUT_DIR}/SOLUCION_vinculacion_red_social.csv")
df_a_hoja(wb, df_sol, "SOLUCION_profesor", max_col_width=35)

# Guardar
out_path = f"{OUT_DIR}/datos_curso_privacidad.xlsx"
print(f"Guardando {out_path}...")
wb.save(out_path)
print("Hecho.")
